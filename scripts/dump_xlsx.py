"""One-off: dump sheet names and row 1 headers from ERP xlsx."""
import glob
import json
import re
import zipfile

from openpyxl import load_workbook

ROOT = r"c:/Users/BernardoAndrade/Documents/Code/EspacoLounge"
paths = glob.glob(f"{ROOT}/*.xlsx")
if not paths:
    raise SystemExit("No xlsx found")
path = paths[0]

# Exact sheet names from OOXML
with zipfile.ZipFile(path, "r") as z:
    xml = z.read("xl/workbook.xml").decode("utf-8")
sheet_names_xml = re.findall(r'sheet name="([^"]+)"', xml)

wb = load_workbook(path, read_only=False, data_only=True)
out = {
    "file": path,
    "sheet_names_workbook_xml": sheet_names_xml,
    "sheet_names_openpyxl": wb.sheetnames,
    "sheets": [],
}
for name in wb.sheetnames:
    ws = wb[name]
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    headers = []
    if max_row >= 1 and max_col >= 1:
        headers = [
            str(ws.cell(1, c).value or "").strip() for c in range(1, max_col + 1)
        ]
    out["sheets"].append(
        {
            "name": name,
            "max_row": max_row,
            "max_column": max_col,
            "headers": headers,
        }
    )
wb.close()
print(json.dumps(out, ensure_ascii=False, indent=2))
