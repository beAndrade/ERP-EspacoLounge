"""
Gera docs/xlsx-manifest.json: abas, cabeçalhos (linha 1), inferência de tipo por amostra.
Executar na raiz: python scripts/xlsx_manifest.py
"""
from __future__ import annotations

import json
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "docs" / "ERP Espaço Lounge.xlsx"
OUT = ROOT / "docs" / "xlsx-manifest.json"
SAMPLE_ROWS = 30


def classify(v) -> str:
    if v is None or v == "":
        return "empty"
    if isinstance(v, bool):
        return "boolean"
    if isinstance(v, int) and not isinstance(v, bool):
        return "integer"
    if isinstance(v, float):
        return "number"
    if isinstance(v, Decimal):
        return "number"
    if isinstance(v, datetime):
        return "datetime"
    if isinstance(v, date):
        return "date"
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return "empty"
        if s.replace(".", "", 1).replace("-", "", 1).isdigit():
            return "text_numeric"
        return "text"
    return "unknown"


def merge_types(types: list[str]) -> str:
    non_empty = [t for t in types if t != "empty"]
    if not non_empty:
        return "empty"
    if all(t == "text" or t == "empty" for t in types):
        return "text"
    if any(t in ("number", "integer", "text_numeric") for t in non_empty):
        if any(t in ("date", "datetime") for t in non_empty):
            return "mixed"
        return "number_like"
    if any(t in ("date", "datetime") for t in non_empty):
        return "date_like"
    return "mixed"


def main():
    if not XLSX.is_file():
        raise SystemExit(f"Ficheiro não encontrado: {XLSX}")

    wb = load_workbook(XLSX, read_only=True, data_only=True)
    sheets_out = []

    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(
            ws.iter_rows(min_row=1, max_row=1 + SAMPLE_ROWS, values_only=True)
        )
        if not rows:
            sheets_out.append(
                {
                    "name": name,
                    "headers": [],
                    "columns": [],
                    "row_count_estimate": 0,
                }
            )
            continue

        header_row = rows[0]
        headers = [
            (str(h).strip() if h is not None else "") for h in (header_row or ())
        ]
        # trim trailing empty header names for width
        while headers and not headers[-1]:
            headers.pop()

        columns = []
        max_col = len(headers)
        for ci in range(max_col):
            h = headers[ci] if ci < len(headers) else ""
            samples = []
            for r in range(1, min(len(rows), 1 + SAMPLE_ROWS)):
                row = rows[r]
                if row and ci < len(row):
                    samples.append(row[ci])
            col_types = [classify(v) for v in samples]
            columns.append(
                {
                    "index": ci + 1,
                    "header_sheet": h,
                    "inferred_type": merge_types(col_types),
                    "non_empty_samples": sum(
                        1 for v in samples if v is not None and str(v).strip() != ""
                    ),
                }
            )

        sheets_out.append(
            {
                "name": name,
                "headers": headers,
                "columns": columns,
                "row_count_estimate": ws.max_row or 0,
            }
        )

    wb.close()

    payload = {
        "source": str(XLSX.relative_to(ROOT)).replace("\\", "/"),
        "sheet_names": [s["name"] for s in sheets_out],
        "sheets": sheets_out,
    }
    OUT.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
